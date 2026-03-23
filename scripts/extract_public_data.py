#!/usr/bin/env python3
"""
extract_public_data.py
神奈川県西部の病院データを病床機能報告(Excel)と医療情報ネット(CSV)から抽出・統合し、
JSON形式で出力するスクリプト。
"""

import json
import re
import csv
from pathlib import Path
import unicodedata
import openpyxl

# ============================================================
# 定数
# ============================================================
PROJECT_DIR = Path(__file__).resolve().parent.parent
EXCEL_PATH = str(PROJECT_DIR / "data/public_data/bed_function_facility.xlsx")
CSV_PATH = str(PROJECT_DIR / "data/public_data/hospital_facility/01-1_hospital_facility_info_20251201.csv")
OUTPUT_PATH = str(PROJECT_DIR / "data/public_data/kanagawa_hospitals_enriched.json")

KANAGAWA_PREF_CODE = "14"

TARGET_CITIES = {
    "小田原市", "平塚市", "秦野市", "伊勢原市", "南足柄市",
    "藤沢市", "茅ヶ崎市", "厚木市", "海老名市",
    "大磯町", "二宮町", "開成町", "松田町", "山北町",
    "箱根町", "中井町", "大井町", "寒川町",
}

AREA_MAPPING = {
    "小田原市": "odawara",
    "箱根町": "odawara",
    "平塚市": "hiratsuka",
    "大磯町": "oiso_ninomiya",
    "二宮町": "oiso_ninomiya",
    "秦野市": "hadano",
    "伊勢原市": "isehara",
    "南足柄市": "minamiashigara_kaisei_oi",
    "開成町": "minamiashigara_kaisei_oi",
    "松田町": "minamiashigara_kaisei_oi",
    "山北町": "minamiashigara_kaisei_oi",
    "中井町": "minamiashigara_kaisei_oi",
    "大井町": "minamiashigara_kaisei_oi",
    "藤沢市": "fujisawa",
    "茅ヶ崎市": "chigasaki",
    "寒川町": "chigasaki",
    "厚木市": "atsugi",
    "海老名市": "ebina",
}

# 設置主体の変換マッピング
OWNER_TYPE_MAP = {
    "都道府県": "公立",
    "市町村": "公立",
    "地方独立行政法人": "公立",
    "日本赤十字社": "公的",
    "済生会": "公的",
    "厚生連": "公的",
    "北海道社会事業協会": "公的",
    "国民健康保険団体連合会": "公的",
    "全国社会保険協会連合会": "公的",
    "社会保険関係団体": "公的",
    "共済組合及びその連合会": "公的",
    "健康保険組合及びその連合会": "公的",
    "国（厚生労働省）": "国立",
    "国（独立行政法人国立病院機構）": "国立",
    "独立行政法人国立病院機構": "国立",
    "国（国立大学法人）": "国立",
    "国（独立行政法人労働者健康安全機構）": "国立",
    "独立行政法人労働者健康安全機構": "国立",
    "国（独立行政法人地域医療機能推進機構）": "国立",
    "独立行政法人地域医療機能推進機構": "国立",
    "国（その他）": "国立",
    "学校法人": "学校法人",
    "医療法人": "医療法人",
    "社会福祉法人": "社会福祉法人",
    "医療生協": "医療生協",
    "会社": "会社",
    "その他の法人": "その他",
    "個人": "個人",
}


# ============================================================
# ユーティリティ
# ============================================================
def zen_to_han(s):
    """全角数字・英字を半角に変換"""
    if s is None:
        return None
    return unicodedata.normalize("NFKC", str(s))


def safe_int(val):
    """値をintに変換。変換できない場合は0を返す。"""
    if val is None:
        return 0
    s = zen_to_han(str(val)).strip()
    if s in ("", "*", "-", "－", "―", "ー", "＊"):
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def safe_float(val):
    """値をfloatに変換。変換できない場合は0.0を返す。"""
    if val is None:
        return 0.0
    s = zen_to_han(str(val)).strip()
    if s in ("", "*", "-", "－", "―", "ー", "＊"):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def clean_name(name):
    """病院名の正規化（マッチング用）"""
    if not name:
        return ""
    s = zen_to_han(str(name))
    # スペース・全角スペースを除去
    s = re.sub(r"[\s\u3000]+", "", s)
    # 表記ゆれの統一: ヶ/ケ/ケ/ヵ → ヶ
    s = s.replace("ケ", "ヶ").replace("ケ", "ヶ").replace("ヵ", "ヶ")
    return s


def extract_facility_suffix(name):
    """施設名のサフィックス部分（例:「山内病院」「ピースハウス病院」）を抽出。
    法人格や財団名を含まない、最も短い施設名を返す。"""
    if not name:
        return ""
    s = clean_name(name)
    # 「〜病院」「〜医院」等の末尾パターンを探す
    suffixes = ["病院", "クリニック", "医院", "診療所", "メディカルセンター", "センター"]
    for suffix in suffixes:
        idx = s.rfind(suffix)
        if idx < 0:
            continue
        end = idx + len(suffix)
        # suffix の前の文字列から、施設名コア部分を逆方向に探す
        # 法人名は「会」「団」「法人」等で終わるので、その直後からが施設名
        before = s[:idx]
        # 法人名区切りパターン: 会、団、法人、連、機構 等の直後
        separators = ["会", "団", "法人", "連", "機構", "社"]
        best_start = 0
        for sep in separators:
            sep_idx = before.rfind(sep)
            if sep_idx >= 0 and sep_idx + len(sep) > best_start:
                best_start = sep_idx + len(sep)
        core = s[best_start:end]
        if len(core) >= 3:  # 最低3文字（例: 「A病院」）
            return core
    return s


def extract_core_name(name):
    """法人格を除去した施設コア名を返す（後方互換）"""
    return extract_facility_suffix(name)


def name_match(excel_name, csv_name):
    """施設名の部分一致マッチング"""
    e = clean_name(excel_name)
    c = clean_name(csv_name)
    if not e or not c:
        return False
    # 完全一致
    if e == c:
        return True
    # 短い方が長い方に含まれるか
    if len(e) <= len(c):
        if e in c:
            return True
    else:
        if c in e:
            return True
    # 法人格を除去したコア名でも照合
    e_core = extract_core_name(excel_name)
    c_core = extract_core_name(csv_name)
    if e_core and c_core:
        if e_core == c_core:
            return True
        if len(e_core) >= 3 and len(c_core) >= 3:
            if e_core in c_core or c_core in e_core:
                return True
    return False


def determine_emergency_level(tertiary, secondary, designated):
    """救急レベルを判定"""
    t = zen_to_han(str(tertiary)).strip() if tertiary else ""
    s = zen_to_han(str(secondary)).strip() if secondary else ""
    d = zen_to_han(str(designated)).strip() if designated else ""
    if "有" in t:
        return "三次救急"
    if "有" in s:
        return "二次救急"
    if "有" in d:
        return "救急告示"
    return "なし"


def convert_owner_type(raw):
    """設置主体を変換"""
    if not raw:
        return "不明"
    s = zen_to_han(str(raw)).strip()
    # 完全一致を試行
    if s in OWNER_TYPE_MAP:
        return OWNER_TYPE_MAP[s]
    # 部分一致で検索
    for key, val in OWNER_TYPE_MAP.items():
        if key in s or s in key:
            return val
    return s  # 変換できなければそのまま


def has_flag(val):
    """「有り」「有」等をbool変換"""
    if not val:
        return False
    s = zen_to_han(str(val)).strip()
    return "有" in s


# ============================================================
# Excel読み込み（病床機能報告）
# ============================================================
def load_excel_data():
    """Excelからの神奈川県西部病院データ抽出"""
    print("Loading Excel data...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    hospitals = {}
    row_count = 0
    skip_count = 0

    # 実データは Excel 行7以降 (min_row=7)
    for row in ws.iter_rows(min_row=7, values_only=True):
        vals = list(row)
        row_count += 1

        # 都道府県コード確認
        pref_code = zen_to_han(str(vals[3])).strip() if vals[3] else ""
        if pref_code != KANAGAWA_PREF_CODE:
            continue

        # 市区町村名確認
        city_name = zen_to_han(str(vals[9])).strip() if vals[9] else ""
        if city_name not in TARGET_CITIES:
            skip_count += 1
            continue

        name = zen_to_han(str(vals[2])).strip() if vals[2] else ""
        if not name:
            continue

        medical_code = zen_to_han(str(vals[1])).strip() if vals[1] else ""
        area_id = AREA_MAPPING.get(city_name, "unknown")

        # 救急車受入件数: [154]が年間合計、ただし*の場合は月別[155-166]を合計
        ambulance_total = safe_int(vals[154])
        if ambulance_total == 0:
            ambulance_total = sum(safe_int(vals[i]) for i in range(155, 167))

        # CT台数: [169-172]の合計 (マルチスライス64列以上, 16-64列, 16列未満, その他)
        ct_count = sum(safe_int(vals[i]) for i in range(169, 173))

        # MRI台数: [173-175]の合計 (3T以上, 1.5-3T, 1.5T未満)
        mri_count = sum(safe_int(vals[i]) for i in range(173, 176))

        # 職員数（偶数列=常勤、奇数列=非常勤、合算）
        doctor_count = round(safe_float(vals[200]) + safe_float(vals[201]))
        nurse_count = round(safe_float(vals[204]) + safe_float(vals[205]))
        assoc_nurse_count = round(safe_float(vals[206]) + safe_float(vals[207]))
        midwife_count = round(safe_float(vals[210]) + safe_float(vals[211]))
        pt_count = round(safe_float(vals[212]) + safe_float(vals[213]))
        ot_count = round(safe_float(vals[214]) + safe_float(vals[215]))
        st_count = round(safe_float(vals[216]) + safe_float(vals[217]))
        pharmacist_count = round(safe_float(vals[218]) + safe_float(vals[219]))
        or_nurse_count = round(safe_float(vals[230]) + safe_float(vals[231]))

        hospital = {
            "name": name,
            "fullName": name,
            "medicalCode": medical_code,
            "areaId": area_id,
            "cityName": city_name,
            "address": "",
            "lat": None,
            "lng": None,
            "website": "",
            "ownerType": convert_owner_type(vals[12]),
            "dpcGroup": zen_to_han(str(vals[13])).strip() if vals[13] else "",
            "emergencyLevel": determine_emergency_level(vals[99], vals[100], vals[101]),
            "ambulanceCount": ambulance_total,
            "generalBeds": 0,
            "therapyBeds": 0,
            "mentalBeds": 0,
            "totalBeds": 0,
            "maxUsedBeds": safe_int(vals[167]),
            "minUsedBeds": safe_int(vals[168]),
            "ctCount": ct_count,
            "mriCount": mri_count,
            "hasDischargeUnit": has_flag(vals[187]),
            "doctorCount": doctor_count,
            "nurseCount": nurse_count,
            "associateNurseCount": assoc_nurse_count,
            "midwifeCount": midwife_count,
            "ptCount": pt_count,
            "otCount": ot_count,
            "stCount": st_count,
            "pharmacistCount": pharmacist_count,
            "orNurseCount": or_nurse_count,
            "dataSource": "病床機能報告R5+医療情報ネット2025.12",
        }

        hospitals[name] = hospital

    wb.close()
    print(f"  Excel: {row_count} rows scanned, {len(hospitals)} Kanagawa-West hospitals found (skipped {skip_count} non-target cities)")
    return hospitals


# ============================================================
# CSV読み込み（医療情報ネット）
# ============================================================
def load_csv_data():
    """CSVから神奈川県の病院データを読み込む"""
    print("Loading CSV data...")
    csv_hospitals = []

    with open(CSV_PATH, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header = next(reader)  # ヘッダー行をスキップ

        for row in reader:
            if len(row) < 65:
                continue
            pref_code = row[7].strip()
            if pref_code != KANAGAWA_PREF_CODE:
                continue

            csv_hospitals.append({
                "id": row[0].strip(),
                "name": zen_to_han(row[1].strip()),
                "address": row[9].strip() if len(row) > 9 else "",
                "lat": row[10].strip() if len(row) > 10 else "",
                "lng": row[11].strip() if len(row) > 11 else "",
                "website": row[12].strip() if len(row) > 12 else "",
                "generalBeds": safe_int(row[57]) if len(row) > 57 else 0,
                "therapyBeds": safe_int(row[58]) if len(row) > 58 else 0,
                "mentalBeds": safe_int(row[61]) if len(row) > 61 else 0,
                "totalBeds": safe_int(row[64]) if len(row) > 64 else 0,
            })

    print(f"  CSV: {len(csv_hospitals)} Kanagawa hospitals loaded")
    return csv_hospitals


# ============================================================
# 名寄せ（マッチング）
# ============================================================
def merge_data(excel_hospitals, csv_hospitals):
    """ExcelとCSVデータを名寄せして統合"""
    print("Merging data...")
    matched = 0
    unmatched_names = []

    for name, hosp in excel_hospitals.items():
        best_match = None
        best_score = 0  # マッチの「良さ」（短い方の名前の長さ）

        for csv_hosp in csv_hospitals:
            if name_match(name, csv_hosp["name"]):
                # より良いマッチを選ぶ（コア名の長さが近いものを優先）
                e_clean = clean_name(name)
                c_clean = clean_name(csv_hosp["name"])
                score = min(len(e_clean), len(c_clean))
                if score > best_score:
                    best_score = score
                    best_match = csv_hosp

        if best_match:
            matched += 1
            hosp["fullName"] = best_match["name"] if best_match["name"] else hosp["fullName"]
            hosp["address"] = best_match["address"]
            try:
                lat = float(best_match["lat"]) if best_match["lat"] and best_match["lat"] != "0.0" else None
                lng = float(best_match["lng"]) if best_match["lng"] and best_match["lng"] != "0.0" else None
                hosp["lat"] = lat
                hosp["lng"] = lng
            except (ValueError, TypeError):
                pass
            hosp["website"] = best_match["website"]
            hosp["generalBeds"] = best_match["generalBeds"]
            hosp["therapyBeds"] = best_match["therapyBeds"]
            hosp["mentalBeds"] = best_match["mentalBeds"]
            hosp["totalBeds"] = best_match["totalBeds"]
        else:
            unmatched_names.append(name)

    print(f"  Matched: {matched}/{len(excel_hospitals)}")
    if unmatched_names:
        print(f"  Unmatched ({len(unmatched_names)}):")
        for n in unmatched_names:
            print(f"    - {n}")

    return excel_hospitals


# ============================================================
# メイン
# ============================================================
def main():
    # Step 1: Excel読み込み
    excel_hospitals = load_excel_data()

    # Step 2: CSV読み込み
    csv_hospitals = load_csv_data()

    # Step 3: 名寄せ
    merged = merge_data(excel_hospitals, csv_hospitals)

    # Step 4: JSON出力
    result = sorted(merged.values(), key=lambda h: (h["areaId"], -h["totalBeds"], h["name"]))

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\nOutput: {OUTPUT_PATH}")
    print(f"Total hospitals: {len(result)}")

    # サンプル出力
    print("\n--- Sample (first 3) ---")
    for h in result[:3]:
        print(json.dumps(h, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
