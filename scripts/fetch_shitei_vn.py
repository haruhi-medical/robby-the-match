#!/usr/bin/env python3
"""
厚生労働省 関東信越厚生局「指定訪問看護ステーション一覧」→ D1 facilities テーブル投入

ソース: https://kouseikyoku.mhlw.go.jp/kantoshinetsu/chousa/houmon.html
データ: hshitei_rYYMM.zip（関東信越10都県のXLSXを同梱）
本スクリプトは関東4都県のみ抽出: 11=埼玉 / 12=千葉 / 13=東京 / 14=神奈川

使い方:
  python3 scripts/fetch_shitei_vn.py --dry-run           # JSON + SQL出力のみ（D1触らない）
  python3 scripts/fetch_shitei_vn.py --import             # 本番D1に投入
  python3 scripts/fetch_shitei_vn.py --version r0803      # 特定バージョン指定
"""

import argparse
import io
import json
import os
import re
import subprocess
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from urllib.request import urlopen, Request

PROJECT_ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_ROOT / "data" / "shitei_vn"
DATA_DIR.mkdir(parents=True, exist_ok=True)
JSON_OUTPUT = DATA_DIR / "visit_nurse_st_kanto.json"
SQL_OUTPUT = DATA_DIR / "visit_nurse_st_d1.sql"
WRANGLER_DIR = PROJECT_ROOT / "api"

BASE_URL = "https://kouseikyoku.mhlw.go.jp/kantoshinetsu/hshitei_{version}.zip"
TARGET_PREF_CODES = {
    "11": "埼玉県",
    "12": "千葉県",
    "13": "東京都",
    "14": "神奈川県",
}
UA = "nurse-robby data ingestion (+https://quads-nurse.com)"

PREF_NAME_RE = re.compile(r"(東京都|神奈川県|千葉県|埼玉県)")
# 先頭が「市」「区」で始まる名称（市川市、市原市、千代田区等）も拾えるよう
# 数字・ハイフン以外を最長一致で市区町村末尾まで取る
CITY_RE = re.compile(r"^([^0-9\-ー－\s]+?[市区町村])")
ZIP_RE = re.compile(r"〒\s*[0-9０-９]{3}[-－][0-9０-９]{4}")


def normalize_zenkaku_digits(s: str) -> str:
    if not s:
        return s
    table = str.maketrans("０１２３４５６７８９－", "0123456789-")
    return s.translate(table)


def strip_postal(addr: str) -> str:
    """住所から〒郵便番号部分を除去"""
    if not addr:
        return ""
    return ZIP_RE.sub("", addr).strip()


def extract_city(addr: str) -> str:
    """住所の先頭部分から市区町村名を抽出"""
    if not addr:
        return ""
    # 都道府県を除去してから市区町村を探す
    clean = PREF_NAME_RE.sub("", addr)
    m = CITY_RE.match(clean)
    return m.group(1) if m else ""


def detect_prefecture(addr: str, pref_code: str) -> str:
    if addr:
        m = PREF_NAME_RE.search(addr)
        if m:
            return m.group(1)
    return TARGET_PREF_CODES.get(pref_code, "")


def download_zip(version: str, cache_path: Path) -> bytes:
    """ZIP をダウンロード（キャッシュあり）"""
    if cache_path.exists():
        return cache_path.read_bytes()
    url = BASE_URL.format(version=version)
    print(f"  Downloading: {url}")
    req = Request(url, headers={"User-Agent": UA})
    with urlopen(req, timeout=30) as resp:
        data = resp.read()
    cache_path.write_bytes(data)
    print(f"  Saved: {cache_path} ({len(data)} bytes)")
    return data


def parse_xlsx(xlsx_bytes: bytes, pref_code: str) -> list:
    """Excel → 事業所リスト"""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb.active

    facilities = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not row or not row[0]:
            continue
        # データ行判定: A列が整数変換可能
        try:
            int(str(row[0]).replace(",", ""))
        except (ValueError, TypeError):
            continue

        # B列: 指定番号 (例: "019,000.9")
        shitei_no = str(row[1]).strip() if row[1] else ""
        name = str(row[2]).strip().replace("\u3000", " ") if row[2] else ""
        raw_addr = str(row[3]).strip() if row[3] else ""
        tel = str(row[4]).strip() if row[4] else ""
        manager = str(row[5]).strip().replace("\u3000", " ") if row[5] else ""
        shitei_date = str(row[6]).strip() if row[6] else ""
        corp_name = str(row[7]).strip().replace("\u3000", " ") if row[7] else ""
        representative = str(row[8]).strip().replace("\u3000", " ") if row[8] else ""
        status = str(row[9]).strip() if len(row) > 9 and row[9] else ""

        if not name:
            continue
        if status != "現存":
            continue  # 休止・廃止は除外

        address = strip_postal(raw_addr)
        address = normalize_zenkaku_digits(address)
        city = extract_city(address)
        prefecture = detect_prefecture(address, pref_code)

        facilities.append({
            "shitei_no": shitei_no,
            "name": name,
            "prefecture": prefecture,
            "city": city,
            "address": address,
            "tel": tel,
            "manager": manager,
            "shitei_date": shitei_date,
            "corp_name": corp_name,
            "representative": representative,
            "status": status,
        })

    return facilities


def fetch_all(version: str) -> list:
    """関東4都県分を取得してマージ + (name, address) でdedup"""
    zip_cache = DATA_DIR / f"hshitei_{version}.zip"
    zip_bytes = download_zip(version, zip_cache)

    all_facilities = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        names = z.namelist()
        for pref_code in TARGET_PREF_CODES:
            targets = [n for n in names if "/" in n and n.split("/")[-1].startswith(pref_code)]
            if not targets:
                print(f"  [WARN] pref={pref_code} not found in ZIP")
                continue
            xlsx_name = targets[0]
            with z.open(xlsx_name) as f:
                xlsx_bytes = f.read()
            facs = parse_xlsx(xlsx_bytes, pref_code)
            # (name, address) で重複排除、指定日が最新のものを残す
            seen = {}
            for f in facs:
                key = (f["name"], f["address"])
                if key not in seen or (f["shitei_date"] or "") > (seen[key]["shitei_date"] or ""):
                    seen[key] = f
            deduped = list(seen.values())
            dup = len(facs) - len(deduped)
            dup_note = f" (dedup -{dup})" if dup else ""
            print(f"  {TARGET_PREF_CODES[pref_code]}: {len(deduped)}件（現存のみ）{dup_note}")
            all_facilities.extend(deduped)
    return all_facilities


def save_json(facilities: list):
    payload = {
        "fetched_at": datetime.now().isoformat(),
        "source": "kouseikyoku_kantoshinetsu_hshitei",
        "total": len(facilities),
        "by_prefecture": {},
        "facilities": facilities,
    }
    for f in facilities:
        p = f["prefecture"] or "不明"
        payload["by_prefecture"][p] = payload["by_prefecture"].get(p, 0) + 1
    JSON_OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  JSON: {JSON_OUTPUT} ({JSON_OUTPUT.stat().st_size / 1024:.0f} KB)")


def write_sql(facilities: list):
    """D1 投入用SQL生成（既存 source='kouseikyoku_vn' を削除して置換）"""
    with open(SQL_OUTPUT, "w", encoding="utf-8") as f:
        f.write("DELETE FROM facilities WHERE source = 'kouseikyoku_vn';\n\n")
        for fac in facilities:
            name = fac["name"].replace("'", "''")
            address = fac["address"].replace("'", "''")
            # sub_type='訪問看護'（カルーセル職種名欄に表示される）
            # 指定番号は departments 経由の識別ではなくJSON側に保持（D1には出さない）
            f.write(
                "INSERT INTO facilities (name, category, sub_type, prefecture, city, address, "
                "lat, lng, bed_count, departments, source, last_synced_at) VALUES "
                f"('{name}', '訪問看護ST', '訪問看護', '{fac['prefecture']}', "
                f"'{fac['city']}', '{address}', NULL, NULL, NULL, '訪問看護', "
                "'kouseikyoku_vn', datetime('now'));\n"
            )
    print(f"  SQL: {SQL_OUTPUT} ({SQL_OUTPUT.stat().st_size / 1024:.0f} KB, {len(facilities)} INSERTs)")


def run_d1_import():
    print("\nImporting to D1 (remote)...")
    env = {**os.environ, "CLOUDFLARE_API_TOKEN": ""}
    result = subprocess.run(
        [
            "npx", "wrangler", "d1", "execute", "nurse-robby-db",
            f"--file={SQL_OUTPUT}", "--remote", "--config", "wrangler.toml",
        ],
        cwd=str(WRANGLER_DIR),
        capture_output=True, text=True, env=env,
    )
    print((result.stdout or "")[-500:])
    if result.returncode != 0:
        print(f"[ERROR] {(result.stderr or '')[-500:]}")
        sys.exit(result.returncode)
    print("D1 import complete.")


def parse_args():
    ap = argparse.ArgumentParser()
    ap.add_argument("--version", default="r0803",
                    help="ZIP version suffix (rYYMM form). Default: r0803 (令和8年3月).")
    ap.add_argument("--dry-run", action="store_true", help="JSON+SQLのみ、D1触らない")
    ap.add_argument("--import", dest="do_import", action="store_true",
                    help="D1に反映（要 wrangler 認証）")
    return ap.parse_args()


def main():
    args = parse_args()
    print(f"== 指定訪問看護ST取得: version={args.version} ==")
    facilities = fetch_all(args.version)
    print(f"\n合計 {len(facilities)} 件（現存のみ・関東4都県）")
    save_json(facilities)
    write_sql(facilities)

    if args.do_import and not args.dry_run:
        run_d1_import()
    else:
        print("\n[--dry-run] D1 反映はスキップ。本番反映は --import を付けて実行してください。")


if __name__ == "__main__":
    main()
